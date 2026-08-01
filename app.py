from __future__ import annotations

import base64
import html
import re
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from config import disruptions_for, get_available_instances
from convert_schedules import build_baseline_workbook

st.set_page_config(
    page_title="Machine Schedule Board",
    page_icon="📅",
    layout="wide",
    initial_sidebar_state="collapsed",
)

APP_DIR = Path(__file__).resolve().parent

# The code accepts either the original filenames or the newer "(1)" filenames.
LEFT_CHEVRON_CANDIDATES = [
    APP_DIR / "chevron-left.png",
    APP_DIR / "chevron-left(1).png",
    APP_DIR / "angle-left.png",
]
RIGHT_CHEVRON_CANDIDATES = [
    APP_DIR / "right-chevron.png",
    APP_DIR / "right-chevron(1).png",
    APP_DIR / "angle-right (1).png",
]

DAY_RE = re.compile(r"^Day\s+(\d+)$", re.IGNORECASE)
MACHINE_RE = re.compile(r"^M\d+$", re.IGNORECASE)


def window_count_from_instance(instance_id: str) -> int:
    match = re.match(r"^[^_]+_\d+_(\d+)(?:_|$)", instance_id.strip())
    if not match:
        raise ValueError(
            f"Could not determine the window count from instance ID "
            f"'{instance_id}'."
        )
    count = int(match.group(1))
    if count not in {2, 4}:
        raise ValueError(
            f"Unsupported window count {count}; expected 2 or 4."
        )
    return count


RECOVERY_FILES = {
    "Local Repair": (
        "local repair",
        "patient_local_repair_schedule.csv",
    ),
    "RESTORE": (
        "restore",
        "restore_schedule.csv",
    ),
    "Full Reoptimization": (
        "full reoptimization",
        "pure_full_reoptimization_schedule.csv",
    ),
}


def recovery_schedule_path(
        instance_id: str,
        disruption_id: str,
        strategy: str,
) -> Path:
    folder, filename = RECOVERY_FILES[strategy]
    return (
            APP_DIR
            / "data"
            / instance_id
            / "recovery"
            / disruption_id
            / folder
            / filename
    )


def baseline_schedule_path(instance_id: str) -> Path:
    return (
            APP_DIR
            / "data"
            / instance_id
            / "baseline"
            / "baseline_schedule.csv"
    )


@st.cache_data(show_spinner=False)
def read_workbook(file_bytes: bytes, window_count: int) -> dict[str, Any]:
    """Read daily machine/window schedules from the supplied Excel workbook."""
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)

    day_sheets: list[tuple[int, str]] = []
    for sheet_name in workbook.sheetnames:
        match = DAY_RE.match(sheet_name.strip())
        if match:
            day_sheets.append((int(match.group(1)), sheet_name))
    day_sheets.sort()

    result: dict[str, Any] = {
        "days": [day for day, _ in day_sheets],
        "window_count": window_count,
        "sheets": {},
    }

    for day, sheet_name in day_sheets:
        worksheet = workbook[sheet_name]

        machine_rows: list[tuple[int, str]] = []
        for row in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row, 1).value
            if isinstance(value, str) and MACHINE_RE.match(value.strip()):
                machine_rows.append((row, value.strip().upper()))

        machines: dict[str, dict[int, list[dict[str, str]]]] = {}

        for index, (start_row, machine) in enumerate(machine_rows):
            end_row = (
                machine_rows[index + 1][0] - 1
                if index + 1 < len(machine_rows)
                else worksheet.max_row
            )

            machines[machine] = {
                window: []
                for window in range(1, window_count + 1)
            }

            for window in range(1, window_count + 1):
                column = window + 1

                for row in range(start_row, end_row + 1):
                    cell = worksheet.cell(row, column)
                    if cell.value is None:
                        continue

                    text = str(cell.value).strip()
                    if not text or "EMPTY" in text.upper():
                        continue

                    rgb = (
                        cell.fill.fgColor.rgb
                        if cell.fill.fill_type == "solid"
                        else None
                    )
                    if rgb and len(rgb) == 8:
                        rgb = rgb[-6:]

                    source_color = (
                        f"#{rgb}"
                        if rgb and re.fullmatch(r"[0-9A-Fa-f]{6}", rgb)
                        else "#C6E0B4"
                    )

                    machines[machine][window].append(
                        {"text": text, "source_color": source_color}
                    )

        result["sheets"][day] = {
            "sheet_name": sheet_name,
            "machines": machines,
        }

    return result


@st.cache_data(show_spinner=False)
def read_affected_fractions(csv_path: str, modified_ns: int) -> pd.DataFrame:
    """Load and validate one affected-fractions disruption file."""
    del modified_ns  # Included only to invalidate the cache when the file changes.
    path = Path(csv_path)
    if not path.exists():
        raise FileNotFoundError(f"Affected-fractions file not found: {path}")

    affected = pd.read_csv(path)
    required = {"patient", "fraction", "machine", "day", "window"}
    missing = required.difference(affected.columns)
    if missing:
        raise ValueError(
            "Affected-fractions CSV is missing required columns: "
            + ", ".join(sorted(missing))
        )

    for column in required:
        affected[column] = pd.to_numeric(
            affected[column], errors="raise"
        ).astype(int)

    return affected


def affected_lookup_from_dataframe(
        affected: pd.DataFrame,
) -> set[tuple[int, int, int, int, int]]:
    """Build exact appointment-location keys for disruption highlighting."""
    return {
        (
            int(row.patient),
            int(row.fraction),
            int(row.day),
            int(row.machine),
            int(row.window),
        )
        for row in affected.itertuples(index=False)
    }


@st.cache_data(show_spinner=False)
def read_schedule_csv(csv_path: str, modified_ns: int) -> pd.DataFrame:
    del modified_ns
    path = Path(csv_path)
    if not path.exists():
        raise FileNotFoundError(f"Schedule CSV not found: {path}")

    schedule = pd.read_csv(path)
    required = {
        "patient",
        "fraction",
        "machine",
        "day",
        "window",
        "duration_minutes",
        "priority",
        "is_preferred_machine",
    }
    missing = required.difference(schedule.columns)
    if missing:
        raise ValueError(
            "Schedule CSV is missing required columns: "
            + ", ".join(sorted(missing))
        )

    for column in (
            "patient",
            "fraction",
            "machine",
            "day",
            "window",
            "duration_minutes",
            "priority",
    ):
        schedule[column] = pd.to_numeric(
            schedule[column], errors="raise"
        ).astype(int)

    if schedule["is_preferred_machine"].dtype != bool:
        schedule["is_preferred_machine"] = (
            schedule["is_preferred_machine"]
            .astype(str)
            .str.strip()
            .str.lower()
            .map({"true": True, "false": False, "1": True, "0": False})
        )

    if schedule["is_preferred_machine"].isna().any():
        raise ValueError(
            "Column 'is_preferred_machine' contains unrecognised values."
        )

    return schedule.sort_values(
        ["day", "machine", "window", "patient", "fraction"],
        kind="stable",
    ).reset_index(drop=True)


def priority_source_color(priority: int) -> str:
    return {
        1: "#C6E0B4",
        2: "#FCE4D6",
        3: "#F4CCCC",
    }.get(int(priority), "#C6E0B4")


def location_tuple(row: Any) -> tuple[int, int, int]:
    return (
        int(row.day),
        int(row.machine),
        int(row.window),
    )


def classify_change(
        baseline_location: tuple[int, int, int],
        recovered_location: tuple[int, int, int],
) -> str:
    changed: list[str] = []

    if baseline_location[0] != recovered_location[0]:
        changed.append("day")
    if baseline_location[1] != recovered_location[1]:
        changed.append("machine")
    if baseline_location[2] != recovered_location[2]:
        changed.append("window")

    if not changed:
        return "unchanged"
    if len(changed) > 1:
        return "moved-multiple"
    return f"moved-{changed[0]}"


def dataframe_to_schedule(
        dataframe: pd.DataFrame,
        *,
        window_count: int,
        baseline: pd.DataFrame | None = None,
        affected_keys: set[tuple[int, int]] | None = None,
) -> tuple[dict[str, Any], dict[str, int]]:
    affected_keys = affected_keys or set()

    baseline_by_key: dict[tuple[int, int], Any] = {}
    if baseline is not None:
        baseline_by_key = {
            (int(row.patient), int(row.fraction)): row
            for row in baseline.itertuples(index=False)
        }

    recovered_keys = {
        (int(row.patient), int(row.fraction))
        for row in dataframe.itertuples(index=False)
    }

    cards: list[dict[str, Any]] = []
    metrics = {
        "modified": 0,
        "recovered": 0,
        "unrecovered": 0,
        "unchanged_affected": 0,
        "changed_nonaffected": 0,
        "day_changes": 0,
        "machine_changes": 0,
        "window_changes": 0,
        "unchanged_total": 0,
        "baseline_total": len(baseline_by_key),
    }

    def make_text(row: Any) -> str:
        suffix = "" if bool(row.is_preferred_machine) else "  NP"
        return (
            f"P{int(row.patient):02d}  |  F{int(row.fraction)}{suffix}\n"
            f"{int(row.duration_minutes)} min"
        )

    for row in dataframe.itertuples(index=False):
        key = (int(row.patient), int(row.fraction))
        recovered_location = location_tuple(row)
        baseline_row = baseline_by_key.get(key)
        baseline_location = (
            location_tuple(baseline_row)
            if baseline_row is not None
            else recovered_location
        )
        raw_change = classify_change(baseline_location, recovered_location)
        is_affected = key in affected_keys

        if baseline_row is not None:
            if raw_change == "unchanged":
                metrics["unchanged_total"] += 1
            else:
                metrics["modified"] += 1

                if baseline_location[0] != recovered_location[0]:
                    metrics["day_changes"] += 1
                if baseline_location[1] != recovered_location[1]:
                    metrics["machine_changes"] += 1
                if baseline_location[2] != recovered_location[2]:
                    metrics["window_changes"] += 1

        if is_affected:
            metrics["recovered"] += 1
            if raw_change == "unchanged":
                status = "affected-recovered-same"
                metrics["unchanged_affected"] += 1
            else:
                status = "affected-recovered-moved"
                if baseline_row is not None:
                    cards.append({
                        "patient": int(baseline_row.patient),
                        "fraction": int(baseline_row.fraction),
                        "day": int(baseline_row.day),
                        "machine": int(baseline_row.machine),
                        "window": int(baseline_row.window),
                        "text": make_text(baseline_row),
                        "source_color": priority_source_color(int(baseline_row.priority)),
                        "status": "affected-origin",
                        "raw_change": raw_change,
                        "is_affected": True,
                        "baseline_location": baseline_location,
                        "recovered_location": recovered_location,
                        "is_preferred_machine": bool(baseline_row.is_preferred_machine),
                    })
        else:
            if raw_change == "unchanged":
                status = "unchanged"
            else:
                status = "nonaffected-changed"
                metrics["changed_nonaffected"] += 1

        cards.append({
            "patient": int(row.patient),
            "fraction": int(row.fraction),
            "day": int(row.day),
            "machine": int(row.machine),
            "window": int(row.window),
            "text": make_text(row),
            "source_color": priority_source_color(int(row.priority)),
            "status": status,
            "raw_change": raw_change,
            "is_affected": is_affected,
            "baseline_location": baseline_location,
            "recovered_location": recovered_location,
            "is_preferred_machine": bool(row.is_preferred_machine),
        })

    if baseline is not None:
        for key in sorted(affected_keys - recovered_keys):
            baseline_row = baseline_by_key.get(key)
            if baseline_row is None:
                continue
            metrics["unrecovered"] += 1
            baseline_location = location_tuple(baseline_row)
            cards.append({
                "patient": int(baseline_row.patient),
                "fraction": int(baseline_row.fraction),
                "day": int(baseline_row.day),
                "machine": int(baseline_row.machine),
                "window": int(baseline_row.window),
                "text": make_text(baseline_row),
                "source_color": priority_source_color(int(baseline_row.priority)),
                "status": "unrecovered",
                "raw_change": "unrecovered",
                "is_affected": True,
                "baseline_location": baseline_location,
                "recovered_location": None,
                "is_preferred_machine": bool(baseline_row.is_preferred_machine),
            })

    days = sorted({int(card["day"]) for card in cards})
    machines = sorted({int(card["machine"]) for card in cards})
    result: dict[str, Any] = {
        "days": days,
        "window_count": window_count,
        "sheets": {},
    }
    order = {
        "affected-origin": 0,
        "unrecovered": 1,
        "affected-recovered-same": 2,
        "affected-recovered-moved": 2,
        "nonaffected-changed": 3,
        "unchanged": 4,
    }

    for day in days:
        day_machines = {
            f"M{machine}": {w: [] for w in range(1, window_count + 1)}
            for machine in machines
        }
        for card in cards:
            if int(card["day"]) != day:
                continue
            machine_name = f"M{int(card['machine'])}"
            day_machines.setdefault(
                machine_name,
                {w: [] for w in range(1, window_count + 1)},
            )
            day_machines[machine_name][int(card["window"])].append(card)

        for machine_name in day_machines:
            for window in day_machines[machine_name]:
                day_machines[machine_name][window].sort(
                    key=lambda card: (
                        order.get(str(card.get("status")), 99),
                        int(card["patient"]),
                        int(card["fraction"]),
                    )
                )

        result["sheets"][day] = {
            "sheet_name": f"Day {day:02d}",
            "machines": day_machines,
        }

    return result, metrics


def natural_machine_key(machine: str) -> tuple[int, str]:
    match = re.search(r"\d+", machine)
    return (int(match.group()) if match else 10_000, machine)


def first_existing_path(paths: list[Path]) -> Path | None:
    for path in paths:
        if path.exists():
            return path
    return None


def image_data_uri(path: Path | None) -> str:
    if path is None or not path.exists():
        return ""
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:image/png;base64,{encoded}"


def priority_class(source_color: str) -> str:
    """Map the workbook colors to three clean dashboard priority accents."""
    value = source_color.upper()

    priority_3_colors = {
        "#F4CCCC",
        "#F8D7DA",
        "#F4B6B6",
        "#FFC7CE",
        "#E6B8AF",
    }
    priority_2_colors = {
        "#FCE4D6",
        "#F9CB9C",
        "#FFD9B3",
        "#FFF2CC",
        "#F6B26B",
    }

    if value in priority_3_colors:
        return "priority-3"
    if value in priority_2_colors:
        return "priority-2"
    return "priority-1"


def appointment_parts(raw: str) -> tuple[str, str, str]:
    """Extract patient, fraction, and duration from an appointment cell."""
    lines = [
        line.strip()
        for line in raw.replace("|", " ").splitlines()
        if line.strip()
    ]
    joined = " ".join(lines)

    patient_match = re.search(r"\bP\s*0*(\d+)\b", joined, re.IGNORECASE)
    fraction_match = re.search(r"\bF\s*0*(\d+)\b", joined, re.IGNORECASE)
    duration_match = re.search(
        r"\b(\d+)\s*(?:m|min|mins|minute|minutes)\b",
        joined,
        re.IGNORECASE,
    )

    patient = (
        f"P{patient_match.group(1)}"
        if patient_match
        else (lines[0] if lines else "Appointment")
    )
    fraction = f"F{fraction_match.group(1)}" if fraction_match else "Not specified"
    duration = f"{duration_match.group(1)} min" if duration_match else "Not specified"

    return patient, fraction, duration


def appointment_card_html(
        card: dict[str, Any],
        *,
        day: int,
        machine: str,
        window: int,
        affected_lookup: set[tuple[int, int, int, int, int]],
        disruption_day: int | None = None,
        disruption_window: int | None = None,
        changes_only: bool = False,
) -> str:
    patient, fraction, duration = appointment_parts(card["text"])
    is_non_preferred = (
            not bool(card.get("is_preferred_machine", True))
            or "NP" in card["text"].upper()
    )

    patient_match = re.search(r"\d+", patient)
    fraction_match = re.search(r"\d+", fraction)
    machine_match = re.search(r"\d+", machine)
    patient_number = int(patient_match.group()) if patient_match else None
    fraction_number = int(fraction_match.group()) if fraction_match else None
    machine_number = int(machine_match.group()) if machine_match else None

    exact_affected = (
            patient_number is not None
            and fraction_number is not None
            and machine_number is not None
            and (
                patient_number,
                fraction_number,
                int(day),
                machine_number,
                int(window),
            ) in affected_lookup
    )

    status = str(card.get("status", "unchanged"))
    raw_change = str(card.get("raw_change", "unchanged"))
    is_before_disruption = (
            disruption_day is not None
            and disruption_window is not None
            and (
                    int(day) < disruption_day
                    or (int(day) == disruption_day and int(window) < disruption_window)
            )
    )

    classes = ["appt-card", priority_class(str(card["source_color"]))]

    if exact_affected and status == "unchanged":
        classes.append("affected-card")

    # The original affected position is always shown as historical context.
    if status == "affected-origin":
        classes.append("affected-origin")

    # Toggle only the recovery comparison overlays.
    if changes_only:
        if status in {"affected-recovered-same", "affected-recovered-moved"}:
            classes.append("recovered-affected")
        elif status == "nonaffected-changed":
            classes.append("changed-nonaffected")
        elif status == "unrecovered":
            classes.append("unrecovered-card")

    # Historical appointments remain dimmed independently of comparison borders.
    if is_before_disruption:
        classes.append("pre-disruption")

    baseline_location = card.get("baseline_location")
    recovered_location = card.get("recovered_location")
    tooltip_lines = [
        f"Patient: {patient}",
        f"Fraction: {fraction}",
        f"Duration: {duration}",
    ]
    if baseline_location:
        tooltip_lines.append(
            f"Baseline: Day {baseline_location[0]} · "
            f"M{baseline_location[1]} · W{baseline_location[2]}"
        )
    if recovered_location:
        tooltip_lines.append(
            f"Recovered: Day {recovered_location[0]} · "
            f"M{recovered_location[1]} · W{recovered_location[2]}"
        )

    labels = {
        "affected-origin": "Affected original position",
        "affected-recovered-same": "Recovered in same position",
        "affected-recovered-moved": "Affected and recovered elsewhere",
        "nonaffected-changed": "Changed although not affected",
        "unrecovered": "Unrecovered",
    }
    if status in labels:
        tooltip_lines.append(f"Status: {labels[status]}")
    if raw_change not in {"unchanged", "unrecovered"}:
        tooltip_lines.append(
            "Change: " + raw_change.replace("-", " ").title()
        )

    tooltip = html.escape("\n".join(tooltip_lines), quote=True)
    patient_text = html.escape(patient)

    badge = (
        '<span class="np-badge">NP</span>'
        if is_non_preferred and status != "unrecovered"
        else ""
    )

    return (
        f'<div class="{" ".join(classes)}" title="{tooltip}">'
        f'{badge}<span>{patient_text}</span></div>'
    )


def build_board(
        day_data: dict[str, Any],
        selected_day: int,
        total_days: int,
        window_count: int,
        affected_lookup: set[tuple[int, int, int, int, int]] | None = None,
        disruption_day: int | None = None,
        disruption_window: int | None = None,
        changes_only: bool = False,
        show_recovery_legend: bool = False,
) -> tuple[str, int]:
    affected_lookup = affected_lookup or set()
    machines = sorted(day_data["machines"], key=natural_machine_key)

    total_appointments = sum(
        len(day_data["machines"][machine][window])
        for machine in machines
        for window in range(1, window_count + 1)
    )
    occupied_windows = sum(
        1
        for machine in machines
        for window in range(1, window_count + 1)
        if day_data["machines"][machine][window]
    )

    machine_headers = "".join(
        (
            f'<th><div class="machine-name">{html.escape(machine)}</div>'
            f'<div class="machine-count">'
            f'{sum(len(day_data["machines"][machine][w]) for w in range(1, window_count + 1))} appts'
            f'</div></th>'
        )
        for machine in machines
    )

    rows: list[str] = []
    total_rows_height = 0

    for window in range(1, window_count + 1):
        maximum_cards = max(
            (
                len(day_data["machines"][machine][window])
                for machine in machines
            ),
            default=0,
        )

        # Two cards per row. Compact enough to keep the page short, but readable.
        card_rows = max(1, (maximum_cards + 1) // 2)
        row_height = max(88, 18 + card_rows * 36)
        total_rows_height += row_height + 6

        cells: list[str] = []
        for machine in machines:
            cards = day_data["machines"][machine][window]

            if cards:
                content = "".join(
                    appointment_card_html(
                        card,
                        day=selected_day,
                        machine=machine,
                        window=window,
                        affected_lookup=affected_lookup,
                        disruption_day=disruption_day,
                        disruption_window=disruption_window,
                        changes_only=changes_only,
                    )
                    for card in cards
                )
                slot_content = f'<div class="slot-grid">{content}</div>'
            else:
                slot_content = '<div class="empty-slot">Available</div>'

            cells.append(
                f'<td style="height:{row_height}px">{slot_content}</td>'
            )

        rows.append(
            f'<tr><th class="window-label">'
            f'<span>W{window}</span></th>'
            f'{"".join(cells)}</tr>'
        )

    # Recovery status is relevant only in recovery strategy views.
    recovery_legend_html = ""
    if show_recovery_legend:
        recovery_legend_html = """
          <span class="priority-legend-title">Recovery</span>
          <span class="priority-legend-item">
            <i class="priority-legend-swatch"
               style="background:#F3F4F6;
                      border:1px solid #E5E7EB;
                      border-left:5px solid #DC2626;">
            </i>
            Disrupted slot
          </span>
          <span class="priority-legend-item">
            <i class="priority-legend-swatch"
               style="background:#FFFFFF;
                      border:4px solid #16A34A;">
            </i>
            Affected recovered
          </span>
          <span class="priority-legend-item">
            <i class="priority-legend-swatch"
               style="background:#FFFFFF;
                      border:4px solid #EA580C;">
            </i>
            Non-affected modified
          </span>
          <span class="priority-legend-item">
            <i class="priority-legend-swatch"
               style="background:#111827;
                      border:4px solid #000000;">
            </i>
            Unrecovered
          </span>
        """

    # Extra space is reserved for the compact priority legend above the table.
    component_height = 230 + total_rows_height

    board_html = f"""
    <style>
      :root {{ color-scheme: light; }}

      html, body {{
        margin: 0;
        padding: 0;
        overflow: visible !important;
      }}

      body {{
        font-family: Inter, ui-sans-serif, system-ui, -apple-system,
                     BlinkMacSystemFont, "Segoe UI", sans-serif;
        background: #FFFFFF;
      }}

      .summary {{
        display: grid;
        grid-template-columns: repeat(3, minmax(150px, 210px));
        justify-content: center;
        gap: 14px;
        width: 100%;
        margin: 0 auto 16px;
        padding: 0;
        background: transparent;
      }}

      .metric {{
        position: relative;
        min-height: 72px;
        box-sizing: border-box;
        display: flex;
        align-items: center;
        justify-content: center;
        padding: 13px 18px;
        overflow: hidden;
        background: #FFFFFF;
        border: 1px solid #DCE4EC;
        border-radius: 14px;
        box-shadow: 0 5px 16px rgba(36, 52, 71, .06);
      }}

      .metric::before {{
        content: "";
        position: absolute;
        inset: 0 auto 0 0;
        width: 4px;
        background: #7D91A5;
      }}

      .metric-day::before {{ background: #C4DADE; }}
      .metric-appts::before {{ background: #B9D9D5; }}
      .metric-machines::before {{ background: #E7D9CC; }}

      .metric-copy {{
        min-width: 0;
        text-align: center;
      }}

      .metric strong {{
        display: block;
        color: #243447;
        font-size: 22px;
        font-weight: 750;
        line-height: 1.05;
        letter-spacing: -.02em;
      }}

      .metric span {{
        display: block;
        margin-top: 5px;
        color: #78879A;
        font-size: 10px;
        font-weight: 600;
        letter-spacing: .09em;
        text-transform: uppercase;
      }}

      .board-wrap {{
        width: 100%;
        overflow: visible;
        padding-bottom: 8px;
      }}

      table {{
        width: 100%;
        table-layout: fixed;
        border-collapse: separate;
        border-spacing: 5px;
      }}

      thead th {{
        min-width: 0;
        padding: 10px 5px;
        color: #FFFFFF;
        background: #34495E;
        border-radius: 10px;
        box-shadow: 0 3px 9px rgba(52, 73, 94, .12);
      }}

      thead th:first-child {{
        width: 76px;
        min-width: 76px;
        background: #293B4D;
      }}

      .machine-name {{
        font-size: 15px;
        font-weight: 700;
      }}

      .machine-count {{
        margin-top: 2px;
        font-size: 9px;
        font-weight: 500;
        opacity: .72;
      }}

      tbody td {{
        min-width: 0;
        padding: 6px;
        vertical-align: top;
        background: #FFFFFF;
        border: 1px solid #E4E9F0;
        border-radius: 10px;
        box-shadow: 0 2px 7px rgba(36, 52, 71, .035);
      }}

      .window-label {{
        width: 76px;
        min-width: 76px;
        padding: 8px 4px;
        text-align: center;
        color: #34495E;
        background: #F1F5F8;
        border: 1px solid #DDE5EC;
        border-radius: 10px;
      }}

      .window-label span {{
        display: block;
        font-size: 21px;
        font-weight: 700;
      }}


      .slot-grid {{
        display: grid;
        grid-template-columns: repeat(2, minmax(0, 1fr));
        align-content: start;
        gap: 5px;
        width: 100%;
        height: 100%;
      }}

      .appt-card {{
        position: relative;
        box-sizing: border-box;
        display: flex;
        align-items: center;
        justify-content: center;
        min-width: 0;
        min-height: 29px;
        padding: 4px 5px;
        overflow: hidden;
        color: #243447;
        background: #ECF2F5;
        border: 1px solid rgba(52, 73, 94, .10);
        border-left-width: 4px;
        border-radius: 7px;
        box-shadow: 0 1px 3px rgba(36, 52, 71, .07);
        cursor: default;
        transition: transform .14s ease, box-shadow .14s ease;
      }}

      .appt-card:hover {{
        transform: translateY(-1px);
        box-shadow: 0 4px 10px rgba(36, 52, 71, .12);
      }}

      .appt-card span:not(.np-badge):not(.affected-badge) {{
        display: block;
        width: 100%;
        overflow: hidden;
        color: #173044;
        font-size: 11px;
        font-weight: 500;
        line-height: 1;
        text-align: center;
        text-overflow: ellipsis;
        white-space: nowrap;
      }}

      /* Full pastel appointment cards using the selected palette. */
      .priority-1 {{
        background: #C3DEDD;
        border-left-color: #82B2C0;
      }}
      .priority-2 {{
        background: #F6C7B3;
        border-left-color: #D89A80;
      }}
      .priority-3 {{
        background: #82B2C0;
        border-left-color: #5C8F9E;
      }}

      .affected-card {{
        background: #FECACA;
        border-color: #FCA5A5;
        border-left-color: #DC2626;
        box-shadow: 0 1px 4px rgba(185, 28, 28, .14);
      }}

      .affected-card span:not(.affected-badge) {{
        color: #7F1D1D;
        font-weight: 650;
      }}

      .pre-disruption {{
        opacity: .32;
        filter: grayscale(.25) saturate(.45);
        pointer-events: none;
      }}

      .dim-unchanged {{
        opacity: 1;
        filter: none;
      }}

      /* Original affected position: retain priority fill, add red border. */
      /* Original affected position (vacated after recovery) */
.affected-origin {{
    background: #F3F4F6 !important;
    color: #6B7280 !important;

    border: 1px solid #E5E7EB !important;
    border-left: 5px solid #DC2626 !important;

    outline: none !important;
    box-shadow: none !important;

    opacity: 0.65;
}}

      /* Affected appointment successfully recovered: green border. */
      .recovered-affected {{
        outline: 4px solid #16A34A;
        outline-offset: -4px;
        box-shadow: 0 1px 7px rgba(22, 163, 74, .28);
      }}

      /* Non-affected appointment changed by recovery: orange border. */
      .changed-nonaffected {{
        outline: 4px solid #EA580C;
        outline-offset: -4px;
        box-shadow: 0 1px 7px rgba(234, 88, 12, .28);
      }}

      .unrecovered-card {{
        color: #FFFFFF !important;
        background: #111827 !important;
        border: 4px solid #000000 !important;
        box-shadow: 0 1px 8px rgba(0, 0, 0, .36);
      }}

      .unrecovered-card span:not(.status-badge) {{
        color: #FFFFFF !important;
      }}

      .status-badge {{
        position: absolute;
        top: 2px;
        right: 2px;
        padding: 1px 4px;
        color: #FFFFFF;
        background: #F28C28;
        border-radius: 999px;
        font-size: 6px;
        font-weight: 700;
      }}

      .status-badge-red {{ background: #DC2626; }}
      .status-badge-green {{ background: #22A447; }}
      .status-badge-black {{
        background: #000000;
        border: 1px solid rgba(255,255,255,.35);
      }}

      .affected-badge {{
        position: absolute;
        top: 2px;
        right: 2px;
        padding: 1px 4px;
        color: #FFFFFF;
        background: #DC2626;
        border-radius: 999px;
        font-size: 6px;
        font-weight: 700;
      }}

      .np-badge {{
        position: absolute;
        top: 2px;
        right: 2px;
        padding: 1px 3px;
        color: #FFFFFF;
        background: #52677A;
        border-radius: 999px;
        font-size: 6px;
        font-weight: 700;
      }}

      .empty-slot {{
        box-sizing: border-box;
        display: flex;
        align-items: center;
        justify-content: center;
        width: 100%;
        height: 100%;
        min-height: 100%;
        color: #A2ADBA;
        background: #FBFCFE;
        border: 1px dashed #CFD8E2;
        border-radius: 8px;
        font-size: 10px;
        font-weight: 400;
      }}

      .priority-legend {{
        display: flex;
        align-items: center;
        justify-content: flex-end;
        gap: 14px;
        min-height: 24px;
        margin: 0 5px 8px;
        color: #526579;
        font-size: 10px;
        font-weight: 600;
      }}

      .priority-legend-title {{
        margin-right: 2px;
        color: #243447;
        font-weight: 750;
      }}

      .priority-legend-item {{
        display: inline-flex;
        align-items: center;
        gap: 6px;
        white-space: nowrap;
      }}

      .priority-legend-swatch {{
        box-sizing: border-box;
        display: inline-block;
        width: 20px;
        height: 13px;
        border: 1px solid rgba(52, 73, 94, .12);
        border-left-width: 4px;
        border-radius: 4px;
        box-shadow: 0 1px 2px rgba(36, 52, 71, .05);
      }}

      .priority-legend-swatch.priority-1 {{
        background: #C3DEDD;
        border-left-color: #82B2C0;
      }}

      .priority-legend-swatch.priority-2 {{
        background: #F6C7B3;
        border-left-color: #D89A80;
      }}

      .priority-legend-swatch.priority-3 {{
        background: #82B2C0;
        border-left-color: #5C8F9E;
      }}

      .affected-legend-swatch {{
        background: #FECACA;
        border-left-color: #DC2626;
      }}

      @media (max-width: 900px) {{
        .priority-legend {{
          justify-content: center;
          gap: 10px;
          font-size: 9px;
        }}
      }}
    </style>

    <div class="summary">
      <div class="metric metric-day">
        <div class="metric-copy">
          <strong>Day {selected_day:02d}</strong>
          <span>of {total_days} days</span>
        </div>
      </div>
      <div class="metric metric-appts">
        <div class="metric-copy">
          <strong>{total_appointments}</strong>
          <span>Appointments</span>
        </div>
      </div>
      <div class="metric metric-machines">
        <div class="metric-copy">
          <strong>{len(machines)}</strong>
          <span>Machines</span>
        </div>
      </div>
    </div>

    <div class="priority-legend" aria-label="Schedule legend">
      <span class="priority-legend-title">Priority</span>
      <span class="priority-legend-item">
        <i class="priority-legend-swatch priority-1"></i>Priority 1
      </span>
      <span class="priority-legend-item">
        <i class="priority-legend-swatch priority-2"></i>Priority 2
      </span>
      <span class="priority-legend-item">
        <i class="priority-legend-swatch priority-3"></i>Priority 3
      </span>
      {recovery_legend_html}
    </div>

    <div class="board-wrap">
      <table>
        <thead>
          <tr><th>Window</th>{machine_headers}</tr>
        </thead>
        <tbody>{''.join(rows)}</tbody>
      </table>
    </div>


    """

    return board_html, component_height


st.markdown(
    """
    <style>
      .block-container {
        max-width: 100% !important;
        width: 100% !important;
        padding-top: .75rem;
        padding-right: .5rem !important;
        padding-bottom: 2rem;
        padding-left: .5rem !important;
      }

      h1 {
        margin-bottom: .1rem;
        color: #243447;
        letter-spacing: -.03em;
      }


      div[data-baseweb="slider"] {
        padding-top: .2rem;
      }

      .side-arrow-wrap {
        display: flex;
        align-items: center;
        justify-content: center;
        width: 100%;
      }

      .side-arrow-link {
        display: flex;
        align-items: center;
        justify-content: center;
        width: 92px;
        height: 190px;
        background: transparent;
        border: 0;
        text-decoration: none !important;
        transition: transform .15s ease, opacity .15s ease;
      }

      .side-arrow-link:hover {
        transform: scale(1.06);
      }

      .side-arrow-link img {
        width: 82px;
        height: 150px;
        object-fit: contain;
        opacity: .78;
        pointer-events: none;
      }

      .side-arrow-link:hover img {
        opacity: .95;
      }

      .side-arrow-disabled {
        opacity: .18;
        pointer-events: none;
      }

      [data-testid="stHeader"] {
        background: transparent;
      }

      .timeline-shell {
        display: grid;
        grid-template-columns: 38px minmax(220px, 980px) 38px;
        align-items: center;
        justify-content: center;
        gap: 8px;
        width: min(1080px, 92%);
        margin: 0 auto 8px;
      }

      .timeline-line {
        position: relative;
        height: 5px;
        overflow: hidden;
        background: #E5E9ED;
        border-radius: 999px;
      }

      .timeline-fill {
        height: 100%;
        background: #82B2C0;
        border-radius: inherit;
      }

      .timeline-link {
        display: flex;
        align-items: center;
        justify-content: center;
        width: 34px;
        height: 34px;
        color: #B8BEC4 !important;
        background: transparent;
        border: 0;
        border-radius: 8px;
        font-size: 25px;
        font-weight: 300;
        line-height: 1;
        text-decoration: none !important;
        transition: color .15s ease, background .15s ease;
      }

      .timeline-link:hover {
        color: #8F989F !important;
        background: #F4F6F8;
      }

      .timeline-disabled {
        color: #E2E6E9 !important;
        pointer-events: none;
      }

      /* Clickable PNG side buttons. Target the first and last board columns
         directly; markdown wrappers cannot reliably contain Streamlit widgets. */
      [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:first-child
      div[data-testid="stButton"] > button,
      [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:last-child
      div[data-testid="stButton"] > button {
        width: 78px !important;
        height: 110px !important;
        padding: 0 !important;
        border: 0 !important;
        border-radius: 0 !important;
        background-color: transparent !important;
        background-repeat: no-repeat !important;
        background-position: center !important;
        background-size: 56px 96px !important;
        box-shadow: none !important;
        color: transparent !important;
        font-size: 0 !important;
        transition: transform .15s ease, opacity .15s ease !important;
      }

      [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:first-child
      div[data-testid="stButton"] > button:hover,
      [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:last-child
      div[data-testid="stButton"] > button:hover {
        transform: scale(1.06);
      }

      [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:first-child
      div[data-testid="stButton"] > button:disabled,
      [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:last-child
      div[data-testid="stButton"] > button:disabled {
        opacity: .18;
      }

      /* Treatment-horizon navigation */
      .st-key-treatment_horizon {
        margin-top: 18px;
      }

      /* This app has only one day slider, so use both keyed and global
         selectors to support different Streamlit DOM versions. */
      .st-key-treatment_horizon [data-testid="stSlider"],
      div[data-testid="stSlider"] {
        padding: 0 !important;
        --primary-color: #82B2C0 !important;
        --primary-color-rgb: 130, 178, 192 !important;
        accent-color: #82B2C0 !important;
      }

      .st-key-treatment_horizon [data-testid="stSlider"] label,
      div[data-testid="stSlider"] label {
        display: none !important;
      }

      .st-key-treatment_horizon div[data-baseweb="slider"],
      div[data-testid="stSlider"] div[data-baseweb="slider"] {
        padding: 2px 0 8px !important;
        --primary-color: #82B2C0 !important;
        --primary-color-rgb: 130, 178, 192 !important;
        accent-color: #82B2C0 !important;
      }

      /* Inactive track */
      .st-key-treatment_horizon div[data-baseweb="slider"] > div,
      div[data-testid="stSlider"] div[data-baseweb="slider"] > div {
        background: #E5E9ED !important;
        background-color: #E5E9ED !important;
      }

      /* Track segments */
      .st-key-treatment_horizon div[data-baseweb="slider"] > div > div,
      div[data-testid="stSlider"] div[data-baseweb="slider"] > div > div {
        height: 5px !important;
        border-radius: 999px !important;
      }

      /* Active track: cover the BaseWeb variants used by Streamlit. */
      .st-key-treatment_horizon div[data-baseweb="slider"] > div > div:first-child,
      .st-key-treatment_horizon div[data-baseweb="slider"] > div > div[style],
      div[data-testid="stSlider"] div[data-baseweb="slider"] > div > div:first-child,
      div[data-testid="stSlider"] div[data-baseweb="slider"] > div > div[style],
      div[data-testid="stSlider"] [role="progressbar"] {
        background: #82B2C0 !important;
        background-color: #82B2C0 !important;
        border-color: #82B2C0 !important;
      }

      /* Thumb */
      .st-key-treatment_horizon div[role="slider"],
      div[data-testid="stSlider"] div[role="slider"] {
        width: 18px !important;
        height: 18px !important;
        background: #82B2C0 !important;
        background-color: #82B2C0 !important;
        border: 3px solid #FFFFFF !important;
        box-shadow: 0 1px 5px rgba(36,52,71,.24) !important;
      }

      /* The floating value above the thumb should use the same accent. */
      .st-key-treatment_horizon [data-testid="stThumbValue"],
      div[data-testid="stSlider"] [data-testid="stThumbValue"] {
        color: #527F8D !important;
        font-weight: 600 !important;
      }

      .treatment-horizon-heading {
        margin: 0 0 6px;
        color: #243447;
        font-size: 13px;
        font-weight: 750;
        letter-spacing: .02em;
      }

      .disruption-summary {
        display: grid;
        grid-template-columns: minmax(220px, 1.35fr) repeat(7, minmax(82px, .72fr));
        align-items: center;
        gap: 10px;
        margin-top: 18px;
        margin-bottom: 10px;
        padding: 10px 14px;
        color: #713838;
        background: #FFF7F7;
        border: 1px solid #F1CCCC;
        border-left: 4px solid #D95C5C;
        border-radius: 12px;
      }

      .disruption-summary-copy strong,
      .disruption-summary-copy span,
      .disruption-stat b,
      .disruption-stat span {
        display: block;
      }

      .disruption-summary-copy strong {
        color: #553131;
        font-size: 14px;
      }

      .disruption-summary-copy span,
      .disruption-stat span {
        margin-top: 2px;
        color: #966D6D;
        font-size: 10px;
        text-transform: uppercase;
        letter-spacing: .05em;
      }

      .disruption-stat {
        display: flex;
        min-height: 58px;
        flex-direction: column;
        justify-content: center;
        padding-left: 14px;
        border-left: 1px solid #F0DADA;
      }

      .disruption-col-1 { grid-column: 2; }
      .disruption-col-2 { grid-column: 3; }
      .disruption-col-3-4 { grid-column: 4 / span 2; }
      .disruption-col-5-6 { grid-column: 6 / span 2; }
      .disruption-col-7 { grid-column: 8; }

      .disruption-stat b {
        color: #6B3333;
        font-size: 13px;
      }

      .recovery-summary-flat {
        display: grid;
        grid-template-columns: minmax(220px, 1.35fr) repeat(7, minmax(82px, .72fr));
        align-items: center;
        gap: 0;
        margin: 0 0 16px;
        padding: 10px 14px;
        background: #F7F8FA;
        border: 1px solid #D7DDE4;
        border-left: 5px solid #A8AFB7;
        border-radius: 13px;
        box-shadow: 0 4px 12px rgba(36, 52, 71, .04);
      }

      .recovery-summary-copy {
        display: flex;
        min-height: 58px;
        flex-direction: column;
        justify-content: center;
        padding: 4px 18px 4px 6px;
      }

      .recovery-summary-copy span {
        color: #7A8491;
        font-size: 9px;
        font-weight: 800;
        letter-spacing: .09em;
        text-transform: uppercase;
      }

      .recovery-summary-copy strong {
        margin-top: 6px;
        color: #243447;
        font-size: 19px;
        font-weight: 800;
        line-height: 1.05;
      }

      .recovery-flat-stat {
        display: flex;
        min-height: 58px;
        flex-direction: column;
        align-items: flex-start;
        justify-content: center;
        padding: 4px 14px;
        border-left: 1px solid #D9DEE5;
      }

      .recovery-flat-stat b {
        color: #243447;
        font-size: 18px;
        font-weight: 500;
        line-height: 1;
      }

      .recovery-flat-stat span {
        margin-top: 7px;
        color: #747F8D;
        font-size: 8px;
        font-weight: 800;
        letter-spacing: .06em;
        text-transform: uppercase;
      }

      .recovery-rate-good {
        color: #16A34A !important;
        font-weight: 900 !important;
      }

      .recovery-rate-bad {
        color: #DC2626 !important;
        font-weight: 900 !important;
      }

      @media (max-width: 1300px) {
        .recovery-summary-flat,
        .disruption-summary {
          grid-template-columns: repeat(4, 1fr);
        }

        .recovery-summary-copy,
        .disruption-summary-copy {
          grid-column: 1 / -1;
          border-bottom: 1px solid #D9DEE5;
          margin-bottom: 4px;
        }

        .recovery-flat-stat,
        .disruption-stat {
          grid-column: auto !important;
          border-left: 0;
          border-right: 1px solid #D9DEE5;
        }
      }

      @media (max-width: 720px) {
        .recovery-summary-flat {
          grid-template-columns: repeat(2, 1fr);
        }
      }

      @media (max-width: 1050px) {
        .disruption-summary {
          grid-template-columns: repeat(3, 1fr);
        }

        .disruption-summary-copy {
          grid-column: 1 / -1;
        }
      }

      @media (max-width: 700px) {
        .recovery-kpi-panel,
        .change-breakdown {
          grid-template-columns: 1fr;
        }
      }

    </style>
    """,
    unsafe_allow_html=True,
)

# Discover all instances automatically from data/<instance_id>/baseline/.
available_instances = get_available_instances()
if not available_instances:
    st.error("No valid instances were found. Add data/<instance_id>/baseline/baseline_schedule.csv.")
    st.stop()

with st.sidebar:
    st.header("Schedule Selection")

    selected_instance = st.selectbox(
        "Instance",
        options=available_instances,
    )

    instance_disruptions = disruptions_for(selected_instance)
    available_disruptions = [
        disruption_id
        for disruption_id, details in instance_disruptions.items()
        if Path(details["affected"]).exists()
        or Path(details["summary"]).exists()
    ]
    if not available_disruptions:
        available_disruptions = list(instance_disruptions)

    selected_disruption = st.selectbox(
        "Disruption",
        options=available_disruptions,
        format_func=lambda disruption_id: (
            f"{disruption_id} — "
            f"{instance_disruptions[disruption_id]['label']}"
        ),
    )

    selected_view = st.selectbox(
        "Schedule View",
        options=[
            "Baseline",
            "Disrupted",
            "Local Repair",
            "RESTORE",
            "Full Reoptimization",
        ],
        index=0,
    )

    changes_only = False
    if selected_view in RECOVERY_FILES:
        changes_only = st.checkbox(
            "Show comparison borders",
            value=True,
            help=(
                "Turn recovery comparison styling on or off. "
                "Appointments before the disruption remain dimmed."
            ),
        )

try:
    selected_window_count = window_count_from_instance(selected_instance)
except ValueError as exc:
    st.error(str(exc))
    st.stop()

try:
    baseline_csv = baseline_schedule_path(selected_instance)
    baseline_df = read_schedule_csv(
        str(baseline_csv.resolve()),
        baseline_csv.stat().st_mtime_ns,
    )

    if selected_view in {"Baseline", "Disrupted"}:
        workbook_path = build_baseline_workbook(selected_instance)
        workbook_bytes = workbook_path.read_bytes()
        workbook_label = str(workbook_path.resolve())
        schedule = read_workbook(
            workbook_bytes,
            selected_window_count,
        )
        recovery_metrics = None
    else:
        recovery_path = recovery_schedule_path(
            selected_instance,
            selected_disruption,
            selected_view,
        )
        recovery_df = read_schedule_csv(
            str(recovery_path.resolve()),
            recovery_path.stat().st_mtime_ns,
        )
        workbook_path = recovery_path
        workbook_bytes = recovery_path.read_bytes()
        workbook_label = str(recovery_path.resolve())
        schedule = None
        recovery_metrics = None
except Exception as exc:
    st.error(f"The selected schedule could not be prepared: {exc}")
    st.info(
        "Check the baseline/recovery CSV paths and rerun the app."
    )
    st.stop()

affected_lookup: set[tuple[int, int, int, int, int]] = set()
affected_keys: set[tuple[int, int]] = set()
affected_df = pd.DataFrame()
affected_day: int | None = None
affected_window: int | None = None

if selected_view != "Baseline":
    affected_path = Path(instance_disruptions[selected_disruption]["affected"])
    try:
        affected_df = read_affected_fractions(
            str(affected_path.resolve()),
            affected_path.stat().st_mtime_ns,
        )
        affected_lookup = affected_lookup_from_dataframe(affected_df)
        affected_keys = {
            (int(row.patient), int(row.fraction))
            for row in affected_df.itertuples(index=False)
        }

        if not affected_df.empty:
            affected_day = int(affected_df["day"].min())
            first_day_rows = affected_df[
                affected_df["day"] == affected_day
                ]
            affected_window = int(first_day_rows["window"].min())
    except Exception as exc:
        st.error(f"The disruption data could not be read: {exc}")
        st.stop()

if selected_view in RECOVERY_FILES:
    try:
        schedule, recovery_metrics = dataframe_to_schedule(
            recovery_df,
            window_count=selected_window_count,
            baseline=baseline_df,
            affected_keys=affected_keys,
        )
    except Exception as exc:
        st.error(f"The recovery comparison could not be built: {exc}")
        st.stop()

view_identity = f"{selected_instance}:{selected_disruption}:{selected_view}:{changes_only}"
if st.session_state.get("schedule_view_identity") != view_identity:
    st.session_state["schedule_view_identity"] = view_identity
    st.session_state.pop("selected_day", None)
    st.session_state.pop("day_slider", None)
    if selected_view != "Baseline" and affected_day is not None:
        st.session_state["selected_day"] = affected_day
        st.session_state["day_slider"] = affected_day

# Reset the selected day whenever a different Excel workbook is loaded.
workbook_identity = (
    f"{workbook_label}:{workbook_path.stat().st_mtime_ns}:"
    f"{len(workbook_bytes)}:{hash(workbook_bytes[:4096])}:"
    f"{selected_disruption}:{selected_view}"
)
if st.session_state.get("workbook_identity") != workbook_identity:
    st.session_state["workbook_identity"] = workbook_identity
    if selected_view != "Baseline" and affected_day is not None:
        st.session_state["selected_day"] = affected_day
        st.session_state["day_slider"] = affected_day
    else:
        st.session_state.pop("selected_day", None)
        st.session_state.pop("day_slider", None)
    st.session_state.pop("bottom_day_timeline", None)
    st.query_params.clear()

if not schedule["days"]:
    st.error("No worksheets named Day 01, Day 02, etc. were found.")
    st.stop()

days = schedule["days"]

if selected_view != "Baseline" and not affected_df.empty:
    affected_patients = int(affected_df["patient"].nunique())
    affected_fractions = int(len(affected_df))
    affected_days = sorted(affected_df["day"].unique().tolist())
    affected_machines = sorted(affected_df["machine"].unique().tolist())
    affected_windows = sorted(affected_df["window"].unique().tolist())

    disruption_label = instance_disruptions[selected_disruption]["label"]
    days_text = ", ".join(map(str, affected_days))
    machines_text = ", ".join(f"M{m}" for m in affected_machines)
    windows_text = ", ".join(f"W{w}" for w in affected_windows)

    st.markdown(
        f"""
        <div class="disruption-summary">
          <div class="disruption-summary-copy">
            <strong>{html.escape(selected_disruption)} — {html.escape(disruption_label)}</strong>
            <span>Selected disruption overlay</span>
          </div>
          <div class="disruption-stat disruption-col-1"><b>{affected_fractions}</b><span>Fractions</span></div>
          <div class="disruption-stat disruption-col-2"><b>{affected_patients}</b><span>Patients</span></div>
          <div class="disruption-stat disruption-col-3-4"><b>{html.escape(days_text)}</b><span>Day</span></div>
          <div class="disruption-stat disruption-col-5-6"><b>{html.escape(machines_text)}</b><span>Machine</span></div>
          <div class="disruption-stat disruption-col-7"><b>{html.escape(windows_text)}</b><span>Window</span></div>
        </div>
        """,
        unsafe_allow_html=True,
    )

if selected_view in RECOVERY_FILES and recovery_metrics is not None:
    affected_count = len(affected_keys)
    recovered_count = recovery_metrics["recovered"]
    recovery_rate = (
        100.0 * recovered_count / affected_count
        if affected_count
        else 0.0
    )

    st.markdown(
        f"""
        <div class="recovery-summary-flat">
          <div class="recovery-summary-copy">
            <span>Recovery method</span>
            <strong>{html.escape(selected_view)}</strong>
          </div>

          <div class="recovery-flat-stat">
            <b>{recovered_count}</b>
            <span>Recovered</span>
          </div>

          <div class="recovery-flat-stat">
            <b>{recovery_metrics['unrecovered']}</b>
            <span>Not recovered</span>
          </div>

          <div class="recovery-flat-stat recovery-rate-stat">
            <b class="{'recovery-rate-good' if recovery_rate == 100 else 'recovery-rate-bad'}">
              {f'{recovery_rate:.0f}%' if recovery_rate.is_integer() else f'{recovery_rate:.1f}%'}
            </b>
            <span>Recovery rate</span>
          </div>

          <div class="recovery-flat-stat">
            <b>{recovery_metrics['modified']}</b>
            <span>Modified</span>
          </div>

          <div class="recovery-flat-stat">
            <b>{recovery_metrics['day_changes']}</b>
            <span>Day</span>
          </div>

          <div class="recovery-flat-stat">
            <b>{recovery_metrics['machine_changes']}</b>
            <span>Machine</span>
          </div>

          <div class="recovery-flat-stat">
            <b>{recovery_metrics['window_changes']}</b>
            <span>Window</span>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

# The selected day is kept in session state and controlled by both the
# clickable PNG side buttons and the draggable slider below the board.
if (
        "selected_day" not in st.session_state
        or st.session_state["selected_day"] not in days
):
    st.session_state["selected_day"] = days[0]

selected_day = st.session_state["selected_day"]
current_index = days.index(selected_day)

board_html, board_height = build_board(
    schedule["sheets"][selected_day],
    selected_day,
    len(days),
    window_count=schedule["window_count"],
    affected_lookup=affected_lookup,
    disruption_day=affected_day,
    disruption_window=affected_window,
    changes_only=changes_only,
    show_recovery_legend=selected_view in RECOVERY_FILES,
)

left_chevron_path = first_existing_path(LEFT_CHEVRON_CANDIDATES)
right_chevron_path = first_existing_path(RIGHT_CHEVRON_CANDIDATES)
left_chevron_uri = image_data_uri(left_chevron_path)
right_chevron_uri = image_data_uri(right_chevron_path)

# Give the real Streamlit buttons the uploaded PNGs as backgrounds.
st.markdown(
    f"""
    <style>
      [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:first-child
      div[data-testid="stButton"] > button {{
        background-image: url('{left_chevron_uri}') !important;
      }}
      [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:last-child
      div[data-testid="stButton"] > button {{
        background-image: url('{right_chevron_uri}') !important;
      }}
    </style>
    """,
    unsafe_allow_html=True,
)


def go_previous() -> None:
    index = days.index(st.session_state["selected_day"])
    if index > 0:
        st.session_state["selected_day"] = days[index - 1]
        st.session_state["day_slider"] = days[index - 1]


def go_next() -> None:
    index = days.index(st.session_state["selected_day"])
    if index < len(days) - 1:
        st.session_state["selected_day"] = days[index + 1]
        st.session_state["day_slider"] = days[index + 1]


# Keep day-navigation arrows in every view for a consistent workflow.
arrow_top_space = max(250, int(board_height * 0.52) - 55)
left_column, board_column, right_column = st.columns(
    [0.75, 14.5, 0.75],
    gap="small",
)

with left_column:
    st.markdown(
        f'<div style="height:{arrow_top_space}px"></div>',
        unsafe_allow_html=True,
    )
    st.button(
        "Previous day",
        key="side_previous_day",
        on_click=go_previous,
        disabled=current_index == 0,
        use_container_width=True,
    )

with board_column:
    st.components.v1.html(
        board_html,
        height=board_height,
        scrolling=False,
    )

with right_column:
    st.markdown(
        f'<div style="height:{arrow_top_space}px"></div>',
        unsafe_allow_html=True,
    )
    st.button(
        "Next day",
        key="side_next_day",
        on_click=go_next,
        disabled=current_index == len(days) - 1,
        use_container_width=True,
    )

# Synchronize the slider after side-button clicks or workbook changes.
if (
        "day_slider" not in st.session_state
        or st.session_state["day_slider"] not in days
):
    st.session_state["day_slider"] = st.session_state["selected_day"]

with st.container(key="treatment_horizon"):
    horizon_left, horizon_center, horizon_right = st.columns(
        [0.75, 14.5, 0.75],
        gap="small",
    )

    with horizon_center:
        st.markdown(
            '<div class="treatment-horizon-heading">'
            'Treatment Horizon'
            '</div>',
            unsafe_allow_html=True,
        )
        slider_day = st.select_slider(
            "Schedule day",
            options=days,
            key="day_slider",
            format_func=lambda day: f"Day {day:02d}",
            label_visibility="collapsed",
        )

if slider_day != st.session_state["selected_day"]:
    st.session_state["selected_day"] = slider_day
    st.rerun()