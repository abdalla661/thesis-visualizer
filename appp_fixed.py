from __future__ import annotations

import base64
import html
import re
from io import BytesIO
from pathlib import Path
from typing import Any

import streamlit as st
from openpyxl import load_workbook


st.set_page_config(
    page_title="Machine Schedule Board",
    page_icon="📅",
    layout="wide",
    initial_sidebar_state="collapsed",
)

APP_DIR = Path(__file__).resolve().parent
DEFAULT_FILE = APP_DIR / "C2_18_4_v2_9 base.xlsx"

# The code accepts either the original filenames or the newer "(1)" filenames.
LEFT_CHEVRON_CANDIDATES = [
    APP_DIR / "chevron-left.png",
    APP_DIR / "chevron-left(1).png",
    APP_DIR / "angle-left.png",
]
RIGHT_CHEVRON_CANDIDATES = [
    APP_DIR / "right-chevron.png",
    APP_DIR / "right-chevron(1).png",
    APP_DIR / "angle-right (1).png",
]

DAY_RE = re.compile(r"^Day\s+(\d+)$", re.IGNORECASE)
MACHINE_RE = re.compile(r"^M\d+$", re.IGNORECASE)


@st.cache_data(show_spinner=False)
def read_workbook(file_bytes: bytes) -> dict[str, Any]:
    """Read daily machine/window schedules from the supplied Excel workbook."""
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)

    day_sheets: list[tuple[int, str]] = []
    for sheet_name in workbook.sheetnames:
        match = DAY_RE.match(sheet_name.strip())
        if match:
            day_sheets.append((int(match.group(1)), sheet_name))
    day_sheets.sort()

    result: dict[str, Any] = {
        "days": [day for day, _ in day_sheets],
        "sheets": {},
    }

    for day, sheet_name in day_sheets:
        worksheet = workbook[sheet_name]

        machine_rows: list[tuple[int, str]] = []
        for row in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row, 1).value
            if isinstance(value, str) and MACHINE_RE.match(value.strip()):
                machine_rows.append((row, value.strip().upper()))

        machines: dict[str, dict[int, list[dict[str, str]]]] = {}

        for index, (start_row, machine) in enumerate(machine_rows):
            end_row = (
                machine_rows[index + 1][0] - 1
                if index + 1 < len(machine_rows)
                else worksheet.max_row
            )

            machines[machine] = {1: [], 2: [], 3: [], 4: []}

            for window in range(1, 5):
                column = window + 1

                for row in range(start_row, end_row + 1):
                    cell = worksheet.cell(row, column)
                    if cell.value is None:
                        continue

                    text = str(cell.value).strip()
                    if not text or "EMPTY" in text.upper():
                        continue

                    rgb = (
                        cell.fill.fgColor.rgb
                        if cell.fill.fill_type == "solid"
                        else None
                    )
                    if rgb and len(rgb) == 8:
                        rgb = rgb[-6:]

                    source_color = (
                        f"#{rgb}"
                        if rgb and re.fullmatch(r"[0-9A-Fa-f]{6}", rgb)
                        else "#C6E0B4"
                    )

                    machines[machine][window].append(
                        {"text": text, "source_color": source_color}
                    )

        result["sheets"][day] = {
            "sheet_name": sheet_name,
            "machines": machines,
        }

    return result


def natural_machine_key(machine: str) -> tuple[int, str]:
    match = re.search(r"\d+", machine)
    return (int(match.group()) if match else 10_000, machine)


def first_existing_path(paths: list[Path]) -> Path | None:
    for path in paths:
        if path.exists():
            return path
    return None


def image_data_uri(path: Path | None) -> str:
    if path is None or not path.exists():
        return ""
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:image/png;base64,{encoded}"


def priority_class(source_color: str) -> str:
    """Map the workbook colors to three clean dashboard priority accents."""
    value = source_color.upper()

    priority_3_colors = {
        "#F4CCCC",
        "#F8D7DA",
        "#F4B6B6",
        "#FFC7CE",
        "#E6B8AF",
    }
    priority_2_colors = {
        "#FCE4D6",
        "#F9CB9C",
        "#FFD9B3",
        "#FFF2CC",
        "#F6B26B",
    }

    if value in priority_3_colors:
        return "priority-3"
    if value in priority_2_colors:
        return "priority-2"
    return "priority-1"


def appointment_parts(raw: str) -> tuple[str, str, str]:
    """Extract patient, fraction, and duration from an appointment cell."""
    lines = [
        line.strip()
        for line in raw.replace("|", " ").splitlines()
        if line.strip()
    ]
    joined = " ".join(lines)

    patient_match = re.search(r"\bP\s*0*(\d+)\b", joined, re.IGNORECASE)
    fraction_match = re.search(r"\bF\s*0*(\d+)\b", joined, re.IGNORECASE)
    duration_match = re.search(
        r"\b(\d+)\s*(?:m|min|mins|minute|minutes)\b",
        joined,
        re.IGNORECASE,
    )

    patient = (
        f"P{patient_match.group(1)}"
        if patient_match
        else (lines[0] if lines else "Appointment")
    )
    fraction = f"F{fraction_match.group(1)}" if fraction_match else "Not specified"
    duration = f"{duration_match.group(1)} min" if duration_match else "Not specified"

    return patient, fraction, duration


def appointment_card_html(card: dict[str, str]) -> str:
    patient, fraction, duration = appointment_parts(card["text"])
    is_non_preferred = "NP" in card["text"].upper()

    tooltip = html.escape(
        f"Patient: {patient}\nFraction: {fraction}\nDuration: {duration}",
        quote=True,
    )
    patient_text = html.escape(patient)
    badge = '<span class="np-badge">NP</span>' if is_non_preferred else ""

    return (
        f'<div class="appt-card {priority_class(card["source_color"])}" '
        f'title="{tooltip}">'
        f'{badge}<span>{patient_text}</span></div>'
    )


def build_board(day_data: dict[str, Any], selected_day: int, total_days: int) -> tuple[str, int]:
    machines = sorted(day_data["machines"], key=natural_machine_key)

    total_appointments = sum(
        len(day_data["machines"][machine][window])
        for machine in machines
        for window in range(1, 5)
    )
    occupied_windows = sum(
        1
        for machine in machines
        for window in range(1, 5)
        if day_data["machines"][machine][window]
    )

    machine_headers = "".join(
        (
            f'<th><div class="machine-name">{html.escape(machine)}</div>'
            f'<div class="machine-count">'
            f'{sum(len(day_data["machines"][machine][w]) for w in range(1, 5))} appts'
            f'</div></th>'
        )
        for machine in machines
    )

    rows: list[str] = []
    total_rows_height = 0

    for window in range(1, 5):
        maximum_cards = max(
            (
                len(day_data["machines"][machine][window])
                for machine in machines
            ),
            default=0,
        )

        # Two cards per row. Compact enough to keep the page short, but readable.
        card_rows = max(1, (maximum_cards + 1) // 2)
        row_height = max(88, 18 + card_rows * 36)
        total_rows_height += row_height + 6

        cells: list[str] = []
        for machine in machines:
            cards = day_data["machines"][machine][window]

            if cards:
                content = "".join(appointment_card_html(card) for card in cards)
                slot_content = f'<div class="slot-grid">{content}</div>'
            else:
                slot_content = '<div class="empty-slot">Available</div>'

            cells.append(
                f'<td style="height:{row_height}px">{slot_content}</td>'
            )

        rows.append(
            f'<tr><th class="window-label">'
            f'<span>W{window}</span></th>'
            f'{"".join(cells)}</tr>'
        )

    # Extra space is reserved for the compact priority legend above the table.
    component_height = 228 + total_rows_height

    board_html = f"""
    <style>
      :root {{ color-scheme: light; }}

      html, body {{
        margin: 0;
        padding: 0;
        overflow: visible !important;
      }}

      body {{
        font-family: Inter, ui-sans-serif, system-ui, -apple-system,
                     BlinkMacSystemFont, "Segoe UI", sans-serif;
        background: #FFFFFF;
      }}

      .summary {{
        display: grid;
        grid-template-columns: repeat(3, minmax(150px, 210px));
        justify-content: center;
        gap: 14px;
        width: 100%;
        margin: 0 auto 16px;
        padding: 0;
        background: transparent;
      }}

      .metric {{
        position: relative;
        min-height: 72px;
        box-sizing: border-box;
        display: flex;
        align-items: center;
        justify-content: center;
        padding: 13px 18px;
        overflow: hidden;
        background: #FFFFFF;
        border: 1px solid #DCE4EC;
        border-radius: 14px;
        box-shadow: 0 5px 16px rgba(36, 52, 71, .06);
      }}

      .metric::before {{
        content: "";
        position: absolute;
        inset: 0 auto 0 0;
        width: 4px;
        background: #7D91A5;
      }}

      .metric-day::before {{ background: #C4DADE; }}
      .metric-appts::before {{ background: #B9D9D5; }}
      .metric-machines::before {{ background: #E7D9CC; }}

      .metric-copy {{
        min-width: 0;
        text-align: center;
      }}

      .metric strong {{
        display: block;
        color: #243447;
        font-size: 22px;
        font-weight: 750;
        line-height: 1.05;
        letter-spacing: -.02em;
      }}

      .metric span {{
        display: block;
        margin-top: 5px;
        color: #78879A;
        font-size: 10px;
        font-weight: 600;
        letter-spacing: .09em;
        text-transform: uppercase;
      }}

      .board-wrap {{
        width: 100%;
        overflow: visible;
        padding-bottom: 8px;
      }}

      table {{
        width: 100%;
        table-layout: fixed;
        border-collapse: separate;
        border-spacing: 5px;
      }}

      thead th {{
        min-width: 0;
        padding: 10px 5px;
        color: #FFFFFF;
        background: #34495E;
        border-radius: 10px;
        box-shadow: 0 3px 9px rgba(52, 73, 94, .12);
      }}

      thead th:first-child {{
        width: 76px;
        min-width: 76px;
        background: #293B4D;
      }}

      .machine-name {{
        font-size: 15px;
        font-weight: 700;
      }}

      .machine-count {{
        margin-top: 2px;
        font-size: 9px;
        font-weight: 500;
        opacity: .72;
      }}

      tbody td {{
        min-width: 0;
        padding: 6px;
        vertical-align: top;
        background: #FFFFFF;
        border: 1px solid #E4E9F0;
        border-radius: 10px;
        box-shadow: 0 2px 7px rgba(36, 52, 71, .035);
      }}

      .window-label {{
        width: 76px;
        min-width: 76px;
        padding: 8px 4px;
        text-align: center;
        color: #34495E;
        background: #F1F5F8;
        border: 1px solid #DDE5EC;
        border-radius: 10px;
      }}

      .window-label span {{
        display: block;
        font-size: 21px;
        font-weight: 700;
      }}


      .slot-grid {{
        display: grid;
        grid-template-columns: repeat(2, minmax(0, 1fr));
        align-content: start;
        gap: 5px;
        width: 100%;
        height: 100%;
      }}

      .appt-card {{
        position: relative;
        box-sizing: border-box;
        display: flex;
        align-items: center;
        justify-content: center;
        min-width: 0;
        min-height: 29px;
        padding: 4px 5px;
        overflow: hidden;
        color: #243447;
        background: #ECF2F5;
        border: 1px solid rgba(52, 73, 94, .10);
        border-left-width: 4px;
        border-radius: 7px;
        box-shadow: 0 1px 3px rgba(36, 52, 71, .07);
        cursor: default;
        transition: transform .14s ease, box-shadow .14s ease;
      }}

      .appt-card:hover {{
        transform: translateY(-1px);
        box-shadow: 0 4px 10px rgba(36, 52, 71, .12);
      }}

      .appt-card span:not(.np-badge) {{
        display: block;
        width: 100%;
        overflow: hidden;
        color: #173044;
        font-size: 11px;
        font-weight: 500;
        line-height: 1;
        text-align: center;
        text-overflow: ellipsis;
        white-space: nowrap;
      }}

      /* Full pastel appointment cards using the selected palette. */
      .priority-1 {{
        background: #C3DEDD;
        border-left-color: #82B2C0;
      }}
      .priority-2 {{
        background: #F6C7B3;
        border-left-color: #D89A80;
      }}
      .priority-3 {{
        background: #82B2C0;
        border-left-color: #5C8F9E;
      }}

      .np-badge {{
        position: absolute;
        top: 2px;
        right: 2px;
        padding: 1px 3px;
        color: #FFFFFF;
        background: #52677A;
        border-radius: 999px;
        font-size: 6px;
        font-weight: 700;
      }}

      .empty-slot {{
        box-sizing: border-box;
        display: flex;
        align-items: center;
        justify-content: center;
        width: 100%;
        height: 100%;
        min-height: 100%;
        color: #A2ADBA;
        background: #FBFCFE;
        border: 1px dashed #CFD8E2;
        border-radius: 8px;
        font-size: 10px;
        font-weight: 400;
      }}

      .priority-legend {{
        display: flex;
        align-items: center;
        justify-content: flex-end;
        gap: 14px;
        min-height: 24px;
        margin: 0 5px 8px;
        color: #526579;
        font-size: 10px;
        font-weight: 600;
      }}

      .priority-legend-title {{
        margin-right: 2px;
        color: #243447;
        font-weight: 750;
      }}

      .priority-legend-item {{
        display: inline-flex;
        align-items: center;
        gap: 6px;
        white-space: nowrap;
      }}

      .priority-legend-swatch {{
        box-sizing: border-box;
        display: inline-block;
        width: 20px;
        height: 13px;
        border: 1px solid rgba(52, 73, 94, .12);
        border-left-width: 4px;
        border-radius: 4px;
        box-shadow: 0 1px 2px rgba(36, 52, 71, .05);
      }}

      .priority-legend-swatch.priority-1 {{
        background: #C3DEDD;
        border-left-color: #82B2C0;
      }}

      .priority-legend-swatch.priority-2 {{
        background: #F6C7B3;
        border-left-color: #D89A80;
      }}

      .priority-legend-swatch.priority-3 {{
        background: #82B2C0;
        border-left-color: #5C8F9E;
      }}

      @media (max-width: 900px) {{
        .priority-legend {{
          justify-content: center;
          gap: 10px;
          font-size: 9px;
        }}
      }}
    </style>

    <div class="summary">
      <div class="metric metric-day">
        <div class="metric-copy">
          <strong>Day {selected_day:02d}</strong>
          <span>of {total_days} days</span>
        </div>
      </div>
      <div class="metric metric-appts">
        <div class="metric-copy">
          <strong>{total_appointments}</strong>
          <span>Appointments</span>
        </div>
      </div>
      <div class="metric metric-machines">
        <div class="metric-copy">
          <strong>{len(machines)}</strong>
          <span>Machines</span>
        </div>
      </div>
    </div>

    <div class="priority-legend" aria-label="Patient priority legend">
      <span class="priority-legend-title">Priority</span>
      <span class="priority-legend-item">
        <i class="priority-legend-swatch priority-1"></i>Priority 1
      </span>
      <span class="priority-legend-item">
        <i class="priority-legend-swatch priority-2"></i>Priority 2
      </span>
      <span class="priority-legend-item">
        <i class="priority-legend-swatch priority-3"></i>Priority 3
      </span>
    </div>

    <div class="board-wrap">
      <table>
        <thead>
          <tr><th>Window</th>{machine_headers}</tr>
        </thead>
        <tbody>{''.join(rows)}</tbody>
      </table>
    </div>


    """

    return board_html, component_height


st.markdown(
    """
    <style>
      .block-container {
        max-width: 100% !important;
        width: 100% !important;
        padding-top: .75rem;
        padding-right: .5rem !important;
        padding-bottom: 2rem;
        padding-left: .5rem !important;
      }

      h1 {
        margin-bottom: .1rem;
        color: #243447;
        letter-spacing: -.03em;
      }

      [data-testid="stFileUploader"] {
        padding: 7px 10px;
        background: #F7F9FC;
        border-radius: 12px;
      }

      div[data-baseweb="slider"] {
        padding-top: .2rem;
      }

      .side-arrow-wrap {
        display: flex;
        align-items: center;
        justify-content: center;
        width: 100%;
      }

      .side-arrow-link {
        display: flex;
        align-items: center;
        justify-content: center;
        width: 92px;
        height: 190px;
        background: transparent;
        border: 0;
        text-decoration: none !important;
        transition: transform .15s ease, opacity .15s ease;
      }

      .side-arrow-link:hover {
        transform: scale(1.06);
      }

      .side-arrow-link img {
        width: 82px;
        height: 150px;
        object-fit: contain;
        opacity: .78;
        pointer-events: none;
      }

      .side-arrow-link:hover img {
        opacity: .95;
      }

      .side-arrow-disabled {
        opacity: .18;
        pointer-events: none;
      }

      /* Keep the Excel input visible, compact and reliable. */
      [data-testid="stHeader"] {
        background: transparent;
      }

      [data-testid="stFileUploader"] {
        max-width: 520px;
        margin: 0 auto 10px;
        padding: 8px 12px;
        background: #FFFFFF;
        border: 1px solid #DDE6ED;
        border-radius: 12px;
      }

      [data-testid="stFileUploaderDropzone"] {
        min-height: 64px;
        padding: 9px 12px;
      }

      .timeline-shell {
        display: grid;
        grid-template-columns: 38px minmax(220px, 980px) 38px;
        align-items: center;
        justify-content: center;
        gap: 8px;
        width: min(1080px, 92%);
        margin: 0 auto 8px;
      }

      .timeline-line {
        position: relative;
        height: 5px;
        overflow: hidden;
        background: #E5E9ED;
        border-radius: 999px;
      }

      .timeline-fill {
        height: 100%;
        background: #82B2C0;
        border-radius: inherit;
      }

      .timeline-link {
        display: flex;
        align-items: center;
        justify-content: center;
        width: 34px;
        height: 34px;
        color: #B8BEC4 !important;
        background: transparent;
        border: 0;
        border-radius: 8px;
        font-size: 25px;
        font-weight: 300;
        line-height: 1;
        text-decoration: none !important;
        transition: color .15s ease, background .15s ease;
      }

      .timeline-link:hover {
        color: #8F989F !important;
        background: #F4F6F8;
      }

      .timeline-disabled {
        color: #E2E6E9 !important;
        pointer-events: none;
      }

      /* Clickable PNG side buttons. Target the first and last board columns
         directly; markdown wrappers cannot reliably contain Streamlit widgets. */
      [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:first-child
      div[data-testid="stButton"] > button,
      [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:last-child
      div[data-testid="stButton"] > button {
        width: 78px !important;
        height: 110px !important;
        padding: 0 !important;
        border: 0 !important;
        border-radius: 0 !important;
        background-color: transparent !important;
        background-repeat: no-repeat !important;
        background-position: center !important;
        background-size: 56px 96px !important;
        box-shadow: none !important;
        color: transparent !important;
        font-size: 0 !important;
        transition: transform .15s ease, opacity .15s ease !important;
      }

      [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:first-child
      div[data-testid="stButton"] > button:hover,
      [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:last-child
      div[data-testid="stButton"] > button:hover {
        transform: scale(1.06);
      }

      [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:first-child
      div[data-testid="stButton"] > button:disabled,
      [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:last-child
      div[data-testid="stButton"] > button:disabled {
        opacity: .18;
      }

      /* Real draggable day slider. */
      .day-slider-wrap {
        width: min(980px, 86%);
        margin: 4px auto 0;
      }

      .day-slider-wrap [data-testid="stSlider"] {
        padding: 0 !important;
      }

      .day-slider-wrap [data-testid="stSlider"] label {
        display: none !important;
      }

      .day-slider-wrap div[data-baseweb="slider"] {
        padding: 2px 0 8px !important;
      }

      .day-slider-wrap div[role="slider"] {
        width: 18px !important;
        height: 18px !important;
        background: #82B2C0 !important;
        border: 3px solid #FFFFFF !important;
        box-shadow: 0 1px 5px rgba(36,52,71,.24) !important;
      }

      .day-slider-wrap [data-baseweb="slider"] > div > div {
        height: 5px !important;
        border-radius: 999px !important;
      }

    </style>
    """,
    unsafe_allow_html=True,
)


# Keep a clear Excel input at the top of the page.
uploaded_file = st.file_uploader(
    "Excel schedule",
    type=["xlsx"],
    help="Upload another workbook to replace the default schedule.",
    key="excel_schedule_upload",
)


if uploaded_file is not None:
    workbook_bytes = uploaded_file.getvalue()
    workbook_label = uploaded_file.name
elif DEFAULT_FILE.exists():
    workbook_bytes = DEFAULT_FILE.read_bytes()
    workbook_label = DEFAULT_FILE.name
else:
    st.error(
        "Place Daily_Machine_Schedule_Clear.xlsx beside app.py, "
        "or upload an Excel workbook."
    )
    st.stop()

# Reset the selected day whenever a different Excel workbook is loaded.
workbook_identity = f"{workbook_label}:{len(workbook_bytes)}:{hash(workbook_bytes[:4096])}"
if st.session_state.get("workbook_identity") != workbook_identity:
    st.session_state["workbook_identity"] = workbook_identity
    st.session_state.pop("selected_day", None)
    st.session_state.pop("bottom_day_timeline", None)
    st.query_params.clear()

try:
    schedule = read_workbook(workbook_bytes)
except Exception as exc:
    st.error(f"The workbook could not be read: {exc}")
    st.stop()


if not schedule["days"]:
    st.error("No worksheets named Day 01, Day 02, etc. were found.")
    st.stop()


days = schedule["days"]

# The selected day is kept in session state and controlled by both the
# clickable PNG side buttons and the draggable slider below the board.
if (
    "selected_day" not in st.session_state
    or st.session_state["selected_day"] not in days
):
    st.session_state["selected_day"] = days[0]

selected_day = st.session_state["selected_day"]
current_index = days.index(selected_day)

board_html, board_height = build_board(
    schedule["sheets"][selected_day], selected_day, len(days)
)

left_chevron_path = first_existing_path(LEFT_CHEVRON_CANDIDATES)
right_chevron_path = first_existing_path(RIGHT_CHEVRON_CANDIDATES)
left_chevron_uri = image_data_uri(left_chevron_path)
right_chevron_uri = image_data_uri(right_chevron_path)

# Give the real Streamlit buttons the uploaded PNGs as backgrounds.
st.markdown(
    f"""
    <style>
      [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:first-child
      div[data-testid="stButton"] > button {{
        background-image: url('{left_chevron_uri}') !important;
      }}
      [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:last-child
      div[data-testid="stButton"] > button {{
        background-image: url('{right_chevron_uri}') !important;
      }}
    </style>
    """,
    unsafe_allow_html=True,
)

def go_previous() -> None:
    index = days.index(st.session_state["selected_day"])
    if index > 0:
        st.session_state["selected_day"] = days[index - 1]
        st.session_state["day_slider"] = days[index - 1]


def go_next() -> None:
    index = days.index(st.session_state["selected_day"])
    if index < len(days) - 1:
        st.session_state["selected_day"] = days[index + 1]
        st.session_state["day_slider"] = days[index + 1]


# Position the arrows at the vertical middle of the schedule table.
# The HTML component contains the summary cards above the table, so a spacer is
# used to place each Streamlit button beside the middle schedule rows (W2/W3).
arrow_top_space = max(250, int(board_height * 0.52) - 55)

# Wider schedule area: smaller arrow columns and a larger center board.
left_column, board_column, right_column = st.columns(
    [0.75, 14.5, 0.75],
    gap="small",
)

with left_column:
    st.markdown(
        f'<div style="height:{arrow_top_space}px"></div>',
        unsafe_allow_html=True,
    )
    st.button(
        "Previous day",
        key="side_previous_day",
        on_click=go_previous,
        disabled=current_index == 0,
        use_container_width=True,
    )

with board_column:
    st.components.v1.html(
        board_html,
        height=board_height,
        scrolling=False,
    )

with right_column:
    st.markdown(
        f'<div style="height:{arrow_top_space}px"></div>',
        unsafe_allow_html=True,
    )
    st.button(
        "Next day",
        key="side_next_day",
        on_click=go_next,
        disabled=current_index == len(days) - 1,
        use_container_width=True,
    )

# Synchronize the slider after side-button clicks or workbook changes.
if (
    "day_slider" not in st.session_state
    or st.session_state["day_slider"] not in days
):
    st.session_state["day_slider"] = st.session_state["selected_day"]

st.markdown('<div class="day-slider-wrap">', unsafe_allow_html=True)
slider_day = st.select_slider(
    "Schedule day",
    options=days,
    key="day_slider",
    format_func=lambda day: f"Day {day:02d}",
    label_visibility="collapsed",
)
st.markdown('</div>', unsafe_allow_html=True)

if slider_day != st.session_state["selected_day"]:
    st.session_state["selected_day"] = slider_day
    st.rerun()