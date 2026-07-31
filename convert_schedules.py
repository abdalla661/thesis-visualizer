from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

from config import ROOT_DIR, baseline_paths, ensure_generated_directories


DAY_SHEET_PREFIX = "Day "
MACHINE_COUNT = 10
WINDOW_COUNT = 4
FIRST_WINDOW_COLUMN = 2  # B
MACHINE_LABEL_RE = re.compile(r"^M(\d+)$", re.IGNORECASE)

REQUIRED_COLUMNS = {
    "patient",
    "fraction",
    "machine",
    "day",
    "window",
    "duration_minutes",
    "priority",
    "is_preferred_machine",
}


def _template_candidates(instance_id: str) -> list[Path]:
    filename = f"{instance_id} base.xlsx"
    return [
        ROOT_DIR / "templates" / filename,
        ROOT_DIR / "data" / "template" / filename,
        ROOT_DIR / filename,
    ]


def find_template_workbook(instance_id: str, template_path: str | Path | None = None) -> Path:
    if template_path is not None:
        path = Path(template_path)
        if not path.is_absolute():
            path = ROOT_DIR / path
        if not path.exists():
            raise FileNotFoundError(f"Template workbook not found: {path}")
        return path

    for candidate in _template_candidates(instance_id):
        if candidate.exists():
            return candidate

    searched = "\n".join(f"  - {path}" for path in _template_candidates(instance_id))
    raise FileNotFoundError(
        "Could not find the baseline Excel template. Place it in one of:\n"
        f"{searched}"
    )


def _read_schedule_csv(csv_path: Path) -> pd.DataFrame:
    if not csv_path.exists():
        raise FileNotFoundError(f"Baseline schedule CSV not found: {csv_path}")

    schedule = pd.read_csv(csv_path)
    missing = REQUIRED_COLUMNS.difference(schedule.columns)
    if missing:
        raise ValueError(
            "Baseline schedule CSV is missing required columns: "
            + ", ".join(sorted(missing))
        )

    numeric_columns = [
        "patient",
        "fraction",
        "machine",
        "day",
        "window",
        "duration_minutes",
        "priority",
    ]
    for column in numeric_columns:
        schedule[column] = pd.to_numeric(schedule[column], errors="raise").astype(int)

    preferred = schedule["is_preferred_machine"]
    if preferred.dtype == bool:
        schedule["is_preferred_machine"] = preferred
    else:
        schedule["is_preferred_machine"] = (
            preferred.astype(str)
            .str.strip()
            .str.lower()
            .map({"true": True, "false": False, "1": True, "0": False})
        )

    if schedule["is_preferred_machine"].isna().any():
        raise ValueError("Column 'is_preferred_machine' contains unrecognised values.")
    if not schedule["machine"].between(1, MACHINE_COUNT).all():
        raise ValueError(f"Machine values must be between 1 and {MACHINE_COUNT}.")
    if not schedule["window"].between(1, WINDOW_COUNT).all():
        raise ValueError(f"Window values must be between 1 and {WINDOW_COUNT}.")
    if not schedule["priority"].isin([1, 2, 3]).all():
        raise ValueError("Priority values must be 1, 2, or 3.")

    return schedule.sort_values(
        ["day", "machine", "window", "patient", "fraction"],
        kind="stable",
    ).reset_index(drop=True)


def _apply_style(style, target) -> None:
    """Apply an immutable copied StyleArray to a target cell."""
    target._style = copy(style)


def _detect_machine_blocks(worksheet) -> dict[int, tuple[int, int]]:
    """
    Detect machine row blocks from labels in column A.

    Returns:
        {machine_number: (first_row, last_row)}
    """
    machine_starts: list[tuple[int, int]] = []

    for row in range(1, worksheet.max_row + 1):
        value = worksheet.cell(row=row, column=1).value
        if not isinstance(value, str):
            continue

        match = MACHINE_LABEL_RE.fullmatch(value.strip())
        if match:
            machine_starts.append((int(match.group(1)), row))

    if not machine_starts:
        raise ValueError(
            f"No machine labels were found in worksheet '{worksheet.title}'."
        )

    machine_starts.sort(key=lambda item: item[1])
    blocks: dict[int, tuple[int, int]] = {}

    for index, (machine, start_row) in enumerate(machine_starts):
        if index + 1 < len(machine_starts):
            end_row = machine_starts[index + 1][1] - 1
        else:
            end_row = worksheet.max_row

        if end_row < start_row:
            raise ValueError(
                f"Invalid row block for M{machine} in worksheet "
                f"'{worksheet.title}'."
            )

        blocks[machine] = (start_row, end_row)

    return blocks


def _window_count_for_instance(instance_id: str) -> int:
    """
    Extract the window count from names such as:
      C3_18_2_v2_3 -> 2
      C2_18_4_v2_9 -> 4
    """
    match = re.match(r"^[^_]+_\d+_(\d+)(?:_|$)", instance_id.strip())
    if not match:
        raise ValueError(
            f"Could not determine window count from instance ID "
            f"'{instance_id}'."
        )

    window_count = int(match.group(1))
    if window_count not in {2, 4}:
        raise ValueError(
            f"Unsupported window count {window_count} in instance ID "
            f"'{instance_id}'. Expected 2 or 4."
        )

    return window_count




def _expand_machine_blocks_for_day(
    worksheet,
    day_schedule: pd.DataFrame,
    window_count: int,
) -> dict[int, tuple[int, int]]:
    """Expand template machine blocks when a day needs extra appointment rows.

    Rows are inserted from the bottom machine upward so later machine blocks are
    shifted safely. Machine labels in column A are then re-merged over their
    new block heights.
    """
    machine_blocks = _detect_machine_blocks(worksheet)

    required_rows: dict[int, int] = {}
    for machine, (block_start, block_end) in machine_blocks.items():
        machine_schedule = day_schedule[day_schedule["machine"] == machine]
        required = 1
        for window in range(1, window_count + 1):
            required = max(
                required,
                int((machine_schedule["window"] == window).sum()),
            )
        required_rows[machine] = required

    expansions = {
        machine: required_rows[machine] - (block_end - block_start + 1)
        for machine, (block_start, block_end) in machine_blocks.items()
        if required_rows[machine] > (block_end - block_start + 1)
    }

    if not expansions:
        return machine_blocks

    min_schedule_row = min(start for start, _ in machine_blocks.values())
    max_schedule_row = max(end for _, end in machine_blocks.values())

    # Unmerge all schedule-body ranges before row insertion. openpyxl does not
    # reliably resize merged ranges when inserting rows.
    for merged_range in list(worksheet.merged_cells.ranges):
        if (
            merged_range.max_row >= min_schedule_row
            and merged_range.min_row <= max_schedule_row
            and merged_range.min_col <= FIRST_WINDOW_COLUMN + 3
        ):
            worksheet.unmerge_cells(str(merged_range))

    # Insert from bottom to top so previously calculated start rows remain valid.
    for machine, extra_rows in sorted(
        expansions.items(),
        key=lambda item: machine_blocks[item[0]][0],
        reverse=True,
    ):
        block_start, block_end = machine_blocks[machine]
        insert_at = block_end + 1
        source_row = block_end
        source_height = worksheet.row_dimensions[source_row].height

        worksheet.insert_rows(insert_at, amount=extra_rows)

        for new_row in range(insert_at, insert_at + extra_rows):
            worksheet.row_dimensions[new_row].height = source_height
            for column in range(1, worksheet.max_column + 1):
                source_cell = worksheet.cell(source_row, column)
                target_cell = worksheet.cell(new_row, column)
                target_cell._style = copy(source_cell._style)
                if source_cell.has_style:
                    target_cell.number_format = source_cell.number_format
                target_cell.value = None

    expanded_blocks = _detect_machine_blocks(worksheet)

    # Restore the machine-label merges over the expanded row blocks.
    for machine, (block_start, block_end) in expanded_blocks.items():
        label_cell = worksheet.cell(block_start, 1)
        if not label_cell.value:
            label_cell.value = f"M{machine}"
        worksheet.merge_cells(
            start_row=block_start,
            start_column=1,
            end_row=block_end,
            end_column=1,
        )

    return expanded_blocks

def _find_style_cells(workbook, window_count: int) -> dict[str, Any]:
    """Find reusable appointment, empty, and neutral slot styles."""
    styles: dict[str, Any] = {}
    neutral_by_offset: dict[int, Any] = {}

    fill_to_key = {
        "00C6E0B4": "priority_1",
        "00FCE4D6": "priority_2",
        "00F4CCCC": "priority_3",
        "00E7E6E6": "empty",
    }
    appointment_fill_colors = set(fill_to_key)

    for sheet_name in workbook.sheetnames:
        if not sheet_name.startswith(DAY_SHEET_PREFIX):
            continue

        worksheet = workbook[sheet_name]
        machine_blocks = _detect_machine_blocks(worksheet)

        for _, (block_start, block_end) in machine_blocks.items():
            for row in range(block_start, block_end + 1):
                offset = row - block_start

                for column in range(
                    FIRST_WINDOW_COLUMN,
                    FIRST_WINDOW_COLUMN + window_count,
                ):
                    cell = worksheet.cell(row=row, column=column)
                    rgb = cell.fill.fgColor.rgb

                    key = fill_to_key.get(rgb)
                    if key and key not in styles:
                        styles[key] = copy(cell._style)

                    if (
                        cell.fill.fill_type == "solid"
                        and rgb
                        and rgb not in appointment_fill_colors
                        and offset not in neutral_by_offset
                    ):
                        neutral_by_offset[offset] = copy(cell._style)

        if len(styles) == 4 and neutral_by_offset:
            break

    missing = {"priority_1", "priority_2", "priority_3", "empty"}.difference(styles)
    if missing:
        raise ValueError(
            "The template does not contain all required appointment styles: "
            + ", ".join(sorted(missing))
        )

    if not neutral_by_offset:
        raise ValueError(
            "The template does not contain a neutral schedule-cell style."
        )

    styles["neutral_by_offset"] = neutral_by_offset
    styles["neutral_fallback"] = next(iter(neutral_by_offset.values()))
    return styles


def _appointment_text(row: pd.Series) -> str:
    patient = int(row["patient"])
    fraction = int(row["fraction"])
    duration = int(row["duration_minutes"])
    suffix = "" if bool(row["is_preferred_machine"]) else "  NP"
    return f"P{patient:02d}  |  F{fraction}{suffix}\n{duration} min"


def _apply_nonpreferred_border(cell) -> None:
    blue = Side(style="medium", color="4472C4")
    cell.border = Border(left=blue, right=blue, top=blue, bottom=blue)


def _clear_day_body(
    worksheet,
    styles: dict[str, Any],
    machine_blocks: dict[int, tuple[int, int]],
    window_count: int,
) -> None:
    """Reset every machine/window schedule cell to a neutral template style."""
    min_schedule_row = min(start for start, _ in machine_blocks.values())
    max_schedule_row = max(end for _, end in machine_blocks.values())

    for merged_range in list(worksheet.merged_cells.ranges):
        if (
            merged_range.min_col >= FIRST_WINDOW_COLUMN
            and merged_range.max_col <= FIRST_WINDOW_COLUMN + window_count - 1
            and merged_range.min_row >= min_schedule_row
            and merged_range.max_row <= max_schedule_row
        ):
            worksheet.unmerge_cells(str(merged_range))

    neutral_by_offset = styles["neutral_by_offset"]
    fallback_neutral = styles["neutral_fallback"]

    for _, (block_start, block_end) in machine_blocks.items():
        for row in range(block_start, block_end + 1):
            offset = row - block_start
            neutral_style = neutral_by_offset.get(offset, fallback_neutral)

            for column in range(
                FIRST_WINDOW_COLUMN,
                FIRST_WINDOW_COLUMN + window_count,
            ):
                cell = worksheet.cell(row=row, column=column)
                cell.value = None
                _apply_style(neutral_style, cell)


def _populate_day_sheet(
    worksheet,
    day_schedule: pd.DataFrame,
    styles: dict[str, Any],
    window_count: int,
) -> None:
    machine_blocks = _expand_machine_blocks_for_day(
        worksheet,
        day_schedule,
        window_count,
    )
    _clear_day_body(
        worksheet,
        styles,
        machine_blocks,
        window_count,
    )

    if not day_schedule.empty:
        day_number = int(day_schedule["day"].iloc[0])
    else:
        day_number = int(worksheet.title.replace(DAY_SHEET_PREFIX, ""))

    worksheet["A1"] = f"DAY {day_number} — MACHINE APPOINTMENT BOARD"

    for machine, (block_start, block_end) in machine_blocks.items():
        rows_available = block_end - block_start + 1
        machine_schedule = day_schedule[
            day_schedule["machine"] == machine
        ]

        max_appointments = 0
        for window in range(1, window_count + 1):
            count = int((machine_schedule["window"] == window).sum())
            max_appointments = max(max_appointments, count)

        if max_appointments > rows_available:
            raise ValueError(
                f"Day {day_number}, machine M{machine} has "
                f"{max_appointments} appointments in one window, but the "
                f"template block supports only {rows_available} rows."
            )

        for offset in range(max_appointments):
            worksheet.row_dimensions[block_start + offset].height = 35

        for window in range(1, window_count + 1):
            column = FIRST_WINDOW_COLUMN + window - 1
            appointments = machine_schedule[
                machine_schedule["window"] == window
            ]

            if appointments.empty:
                worksheet.merge_cells(
                    start_row=block_start,
                    start_column=column,
                    end_row=block_end,
                    end_column=column,
                )
                cell = worksheet.cell(block_start, column)
                _apply_style(styles["empty"], cell)
                cell.value = "— EMPTY —"
                continue

            if len(appointments) > rows_available:
                raise ValueError(
                    f"Day {day_number}, machine M{machine}, window {window} "
                    f"has {len(appointments)} appointments; the template "
                    f"supports at most {rows_available}."
                )

            for offset, (_, appointment) in enumerate(appointments.iterrows()):
                row_number = block_start + offset
                cell = worksheet.cell(row_number, column)
                priority = int(appointment["priority"])
                _apply_style(styles[f"priority_{priority}"], cell)
                cell.value = _appointment_text(appointment)

                if not bool(appointment["is_preferred_machine"]):
                    _apply_nonpreferred_border(cell)


def build_baseline_workbook(
    instance_id: str,
    *,
    force: bool = False,
    template_path: str | Path | None = None,
) -> Path:
    """Build the day-by-day baseline workbook from baseline_schedule.csv."""
    ensure_generated_directories(instance_id)

    baseline = baseline_paths(instance_id)
    source_csv = Path(baseline["schedule_csv"])
    output_excel = Path(baseline["schedule_excel"])
    template = find_template_workbook(instance_id, template_path)

    if output_excel.exists() and not force:
        newest_input = max(source_csv.stat().st_mtime, template.stat().st_mtime)
        if output_excel.stat().st_mtime >= newest_input:
            return output_excel

    schedule = _read_schedule_csv(source_csv)
    window_count = _window_count_for_instance(instance_id)

    if not schedule["window"].between(1, window_count).all():
        invalid_windows = sorted(
            schedule.loc[
                ~schedule["window"].between(1, window_count),
                "window",
            ].unique().tolist()
        )
        raise ValueError(
            f"Schedule contains windows {invalid_windows}, but instance "
            f"'{instance_id}' is configured for {window_count} windows."
        )

    workbook = load_workbook(template)
    styles = _find_style_cells(workbook, window_count)

    available_days = sorted(schedule["day"].unique().tolist())
    for sheet_name in workbook.sheetnames:
        if not sheet_name.startswith(DAY_SHEET_PREFIX):
            continue
        try:
            day = int(sheet_name.replace(DAY_SHEET_PREFIX, ""))
        except ValueError:
            continue
        day_schedule = schedule[schedule["day"] == day]
        _populate_day_sheet(
            workbook[sheet_name],
            day_schedule,
            styles,
            window_count,
        )

    missing_sheets = [
        day for day in available_days if f"Day {day:02d}" not in workbook.sheetnames
    ]
    if missing_sheets:
        raise ValueError(
            "Template is missing required day sheets: "
            + ", ".join(f"Day {day:02d}" for day in missing_sheets)
        )

    output_excel.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_excel)
    return output_excel


def load_baseline_workbook(
    instance_id: str,
    *,
    force_rebuild: bool = False,
    template_path: str | Path | None = None,
):
    workbook_path = build_baseline_workbook(
        instance_id,
        force=force_rebuild,
        template_path=template_path,
    )
    return load_workbook(workbook_path, data_only=False)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("instance_id")
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()
    path = build_baseline_workbook(args.instance_id, force=args.force)
    print(f"Baseline workbook created: {path}")